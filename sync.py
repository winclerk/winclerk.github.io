import os
import json
import base64
import io
import requests
from datetime import datetime, timezone, date
import re
from pypdf import PdfReader
from openpyxl import load_workbook
from permit_notify import notify_status_changes
from permit_pdf import generate_for_rows as generate_permit_pdfs, PUBLIC_PATH as PDF_PUBLIC_PATH


TENANT_ID     = os.environ["AZURE_TENANT_ID"]
CLIENT_ID     = os.environ["AZURE_CLIENT_ID"]
CLIENT_SECRET = os.environ["AZURE_CLIENT_SECRET"]
GH_PAT        = os.environ["GH_PAT"]

GITHUB_REPO = "winclerk/winclerk.github.io"
GITHUB_FILE = "data.json"

SITE_HOSTNAME = "townofwinchester54557.sharepoint.com"
SITE_PATH     = "/sites/TownBoard"
LIBRARY_NAME  = "Documents"

NEXT_MEETING_PATH = "All Town Board Files/Next Meeting"
PREV_REGULAR_PATH = "All Town Board Files/Previous Regular Meetings"
PREV_SPECIAL_PATH = "All Town Board Files/Previous Special Meetings"

SKIP_FOLDER_NAME = "Internal Only"

# Non-meeting sites: flat document libraries, not meeting folders.
# Each library is scanned recursively up to MAX_SCAN_DEPTH subfolder levels.
MAX_SCAN_DEPTH = 1

FLAT_SITES = [
    {
        "key": "boardsCommissions",
        "path": "/sites/BoardsCommitteesCommissions",
        "libraries": [
            "Planning Commission",
            "Intermunicipal Committees",
            "NEMSD Shared Services",
        ],
    },
    {
        "key": "governance",
        "path": "/sites/Governance",
        "libraries": [
            "Ordinances",
            "Resolutions",
            "Notices",
            "Policies",
            "Fee Schedule",
        ],
    },
    {
        "key": "elections",
        "path": "/sites/Elections",
        "libraries": [
            "Election Notices",
            "Canvass Results & Certifications",
            "Voters",
            "Election Inspectors",
        ],
    },
]

ICAL_URL = "https://winchesterwi.com/?post_type=tribe_events&ical=1&eventDisplay=list"
MEETING_KEYWORDS = ["regular town board meeting", "special town board meeting"]

# ── Permits sync config ──────────────────────────────────
PERMITS_GITHUB_FILE = "permits.json"
PERMITS_DRIVE_ID = "b!c95LT9gy6kqiItDGFto7RFnHF7mxITJGsCZJt01CCvi1TQvoZoFtT7YMKBgrUG67"
PERMITS_ITEM_ID  = "01UVO6XCSTSHOVIGOYBJE3ZIQJYCZ3YSTV"
PERMITS_SHEET    = "Permits"
PERMITS_HEADER_ROW   = 2
PERMITS_FIRST_DATA   = 3
PERMITS_PUBLISHED    = {"proposed", "approved"}

GRAPH_BASE = "https://graph.microsoft.com/v1.0"


def get_token():
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    data = {
        "client_id":     CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "grant_type":    "client_credentials",
        "scope":         "https://graph.microsoft.com/.default",
    }
    r = requests.post(url, data=data)
    if not r.ok:
        print(f"Token error: {r.text}")
    r.raise_for_status()
    return r.json()["access_token"]


def graph_get(token, url):
    r = requests.get(url, headers={"Authorization": f"Bearer {token}"})
    r.raise_for_status()
    return r.json()


def get_site_id(token, site_path=SITE_PATH):
    url = f"https://graph.microsoft.com/v1.0/sites/{SITE_HOSTNAME}:{site_path}"
    return graph_get(token, url)["id"]


def get_drive_id(token, site_id, library_name=LIBRARY_NAME):
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
    for d in graph_get(token, url)["value"]:
        if d["name"] == library_name:
            return d["id"]
    raise ValueError(f"Drive '{library_name}' not found")


def list_root(token, drive_id):
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root/children"
    items = []
    while url:
        data = graph_get(token, url)
        items.extend(data.get("value", []))
        url = data.get("@odata.nextLink")
    return items


def list_children(token, drive_id, folder_path):
    encoded = requests.utils.quote(folder_path)
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{encoded}:/children"
    items = []
    while url:
        data = graph_get(token, url)
        items.extend(data.get("value", []))
        url = data.get("@odata.nextLink")
    return items


def get_item(token, drive_id, item_path):
    encoded = requests.utils.quote(item_path)
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{encoded}"
    return graph_get(token, url)


def get_root_item(token, drive_id):
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root"
    return graph_get(token, url)


def make_folder_link(token, drive_id, folder_path=None):
    """Create (or fetch) a public view link for a folder itself, not a file
    inside it. folder_path=None means the library root. Returns None on
    failure (e.g. anonymous sharing blocked for this site) rather than
    raising, so callers can degrade gracefully."""
    try:
        item = get_item(token, drive_id, folder_path) if folder_path else get_root_item(token, drive_id)
        return make_link(token, drive_id, item["id"])
    except Exception as e:
        print(f"    Warning: could not create folder link for '{folder_path or '(root)'}': {e}")
        return None


def make_link(token, drive_id, item_id):
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/createLink"
    r = requests.post(url,
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        json={"type": "view", "scope": "anonymous"})
    r.raise_for_status()
    return r.json()["link"]["webUrl"]


def download_file(token, drive_id, item_id):
    """Download a file's content from SharePoint via Graph API.
    Returns raw bytes, or None on failure."""
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/content"
    try:
        r = requests.get(url, headers={"Authorization": f"Bearer {token}"})
        r.raise_for_status()
        return r.content
    except Exception as e:
        print(f"    Warning: could not download file: {e}")
        return None


def extract_meeting_details_from_pdf(pdf_bytes):
    """Extract meeting time and location from a PDF agenda's header text.

    Looks for patterns like:
      'Wednesday, August 5 at 10:00 AM'
      'Winchester Town Hall • 7228 CTH W, Winchester, WI 54557'

    Returns a dict with 'time' and/or 'location' keys, or empty dict."""
    try:
        reader = PdfReader(io.BytesIO(pdf_bytes))
        # Only need the first page — meeting details are always in the header
        text = reader.pages[0].extract_text() or ""
    except Exception as e:
        print(f"    Warning: could not read PDF: {e}")
        return {}

    result = {}

    # Extract time: look for "at H:MM AM/PM" or standalone "H:MM AM/PM"
    # on a line that also contains a day of week or date context
    time_match = re.search(
        r'(?:at\s+)?(\d{1,2}:\d{2}\s*[AaPp]\.?[Mm]\.?)',
        text
    )
    if time_match:
        raw_time = time_match.group(1).strip()
        # Normalize: remove dots, ensure space before AM/PM
        raw_time = raw_time.replace(".", "")
        raw_time = re.sub(r'(\d)([AaPp])', r'\1 \2', raw_time)
        raw_time = raw_time.upper()
        result["time"] = raw_time

    # Extract location: look for "Town Hall" mention with address
    loc_match = re.search(
        r'(Winchester\s+Town\s+Hall)',
        text,
        re.IGNORECASE
    )
    if loc_match:
        result["location"] = "Winchester Town Hall"

    return result


def extract_details_from_next_meeting(token, drive_id, docs):
    """Try to extract time/location from the first agenda PDF in the Next Meeting folder.
    Returns a dict with 'time' and/or 'location' keys, or empty dict."""
    for doc in docs:
        fn = doc["filename"]
        if not fn.lower().startswith("agenda") or not fn.lower().endswith(".pdf"):
            continue
        # Need the item ID — we'll look it up by path
        try:
            encoded = requests.utils.quote(f"{NEXT_MEETING_PATH}/{fn}")
            url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/root:/{encoded}"
            item = graph_get(token, url)
            pdf_bytes = download_file(token, drive_id, item["id"])
            if pdf_bytes:
                details = extract_meeting_details_from_pdf(pdf_bytes)
                if details:
                    print(f"    Parsed from PDF: {details}")
                    return details
        except Exception as e:
            print(f"    Warning: could not parse {fn}: {e}")
            continue
    return {}


def _fetch_ical_events():
    """Fetch and parse all events from the WordPress iCal feed.
    Returns a list of dicts with summary, dtstart_raw, location."""
    try:
        r = requests.get(ICAL_URL, timeout=10,
            headers={"User-Agent": "Winchester-Sync/1.0"})
        r.raise_for_status()
    except Exception as e:
        print(f"  Warning: could not fetch iCal feed: {e}")
        return []

    events = []
    current = {}
    in_event = False

    for line in r.text.splitlines():
        line = line.strip()
        if line == "BEGIN:VEVENT":
            in_event = True
            current = {}
        elif line == "END:VEVENT":
            if current:
                events.append(current)
            in_event = False
            current = {}
        elif in_event:
            if line.startswith("SUMMARY:"):
                current["summary"] = line[8:].strip()
            elif line.startswith("DTSTART"):
                current["dtstart_raw"] = line.split(":", 1)[-1].strip()
            elif line.startswith("LOCATION:"):
                current["location"] = line[9:].strip()

    return events


def _parse_event_dt(raw):
    """Parse an iCal DTSTART value into a datetime. Returns None on failure."""
    try:
        if "T" in raw:
            raw_clean = re.sub(r"[:-]", "", raw.replace("Z", ""))
            dt = datetime.strptime(raw_clean[:15], "%Y%m%dT%H%M%S")
            return dt.replace(tzinfo=timezone.utc)
        else:
            dt = datetime.strptime(raw[:8], "%Y%m%d")
            return dt.replace(tzinfo=timezone.utc)
    except Exception:
        return None


def _event_to_result(dt, ev):
    """Convert a parsed event into a result dict with title, date, time, location."""
    raw_location = ev.get("location", "") or ""
    location = raw_location.split("\\,")[0].split(",")[0].strip()
    if not location:
        location = "Winchester Town Hall"
    if dt.hour or dt.minute:
        time_str = dt.strftime("%-I:%M %p")
    else:
        time_str = "6:00 PM"
    return {
        "title":    ev.get("summary", ""),
        "date":     dt.strftime("%Y-%m-%d"),
        "time":     time_str,
        "location": location,
    }


def get_next_meeting_from_ical(ical_events=None, match_date=None):
    """Find the next meeting from the iCal feed.

    Strategy:
    1. Try keyword match (MEETING_KEYWORDS) — finds standard regular/special meetings.
    2. If that fails and match_date is provided, find any event on that date.
    This handles non-standard meetings like budget meetings that aren't named
    with the standard keywords but are still in the Events Calendar."""
    if ical_events is None:
        ical_events = _fetch_ical_events()

    now = datetime.now(timezone.utc)

    # Pass 1: keyword match (standard meeting names)
    candidates = []
    for ev in ical_events:
        summary = ev.get("summary", "")
        if not any(kw in summary.lower() for kw in MEETING_KEYWORDS):
            continue
        dt = _parse_event_dt(ev.get("dtstart_raw", ""))
        if dt and dt >= now:
            candidates.append((dt, ev))

    if candidates:
        candidates.sort(key=lambda x: x[0])
        dt, ev = candidates[0]
        return _event_to_result(dt, ev)

    # Pass 2: date match (any event on the meeting date from SharePoint)
    if match_date:
        for ev in ical_events:
            dt = _parse_event_dt(ev.get("dtstart_raw", ""))
            if dt and dt.strftime("%Y-%m-%d") == match_date:
                print(f"  Found calendar event by date match: {ev.get('summary', '')}")
                return _event_to_result(dt, ev)

    print("  Warning: no matching town board meetings found in iCal feed.")
    return None


def clean_name(s):
    s = re.sub(r"_\d{8}$", "", s)
    s = re.sub(r"_\d{6}$", "", s)
    s = s.replace("-", " ").replace("_", " ").strip()
    s = re.sub(r"(\d{4}) (\d{2})\b", r"\1-\2", s)
    return s


def infer_label(filename):
    name = filename
    for ext in [".pdf", ".docx", ".xlsx", ".doc", ".xls", ".pptx"]:
        if name.lower().endswith(ext):
            name = name[:-len(ext)]
            break

    if re.search(r"NEMSD.{0,5}[Ii]ntermunicipal", name):
        return "NEMSD Intermunicipal Agreement - Current Signed Agreement"

    m = re.match(r"Agenda_(RTBM|STBM|TBSM|BOR)_\d{8}", name)
    if m:
        type_map = {"RTBM": "Regular Meeting Agenda", "STBM": "Special Meeting Agenda",
                    "TBSM": "Special Meeting Agenda", "BOR": "Board of Review Agenda"}
        return type_map[m.group(1)]

    m = re.match(r"Agenda_\d{8}", name)
    if m:
        return "Agenda"

    # Agenda with descriptive slug: Agenda_Some-Description_YYYYMMDD
    # e.g. Agenda_2027-Initial-Budget-Meeting_20260805 -> "2027 Initial Budget Meeting Agenda"
    m = re.match(r"Agenda_(.+?)_(\d{8})$", name)
    if m:
        slug = m.group(1).replace("-", " ").replace("_", " ").strip()
        return f"{slug} Agenda"

    m = re.match(r"Minutes_(BOR_)?\d{8}(_DRAFT)?", name)
    if m:
        is_bor = bool(m.group(1))
        is_draft = bool(m.group(2))
        base = "Board of Review Minutes" if is_bor else "Minutes"
        return f"{base} (Draft)" if is_draft else base

    m = re.match(r"Minutes_(?:RTBM_|STBM_|TBSM_)\d{8}(_DRAFT)?", name)
    if m:
        return "Minutes (Draft)" if m.group(1) else "Minutes"

    m = re.match(r"Report_Clerk_\d{8}", name)
    if m:
        return "Clerk's Report"

    m = re.match(r"Clerks?_Report_\d{8}", name, re.IGNORECASE)
    if m:
        return "Clerk's Report"

    m = re.match(r"Report_Treasurer_\d{6}", name)
    if m:
        return "Treasurer's Report"

    m = re.match(r"Report_NEMSD_\d{6}", name)
    if m:
        return "NEMSD Report"

    m = re.match(r"Report_Pedalers_\d{6}", name)
    if m:
        return "Pedalers Report"

    m = re.match(r"(?:Report_)?Disbursements_(\d{6})$", name)
    if m:
        dt = datetime.strptime(m.group(1), "%Y%m")
        return f"{dt.strftime('%B')} Disbursements"

    m = re.match(r"Report_(.+?)_\d{6}$", name)
    if m:
        return f"{m.group(1).replace('-', ' ').replace('_', ' ')} Report"

    m = re.match(r"Report_(.+?)_\d{8}$", name)
    if m:
        return f"{m.group(1).replace('-', ' ').replace('_', ' ')} Report"

    m = re.match(r"Policy_(\d{4}-\d{2})_(.+)", name)
    if m:
        return f"{m.group(1)} {clean_name(m.group(2))} Policy"

    m = re.match(r"Resolution_(\d{4}-\d{2})_(.+)", name)
    if m:
        return f"{m.group(1)} {clean_name(m.group(2))} Resolution"

    m = re.match(r"Ordinance_(\d{4}-\d{2})_(.+)", name)
    if m:
        return f"{m.group(1)} {clean_name(m.group(2))} Ordinance"

    m = re.match(r"Permit_(.+?)(?:_\d{8})?$", name)
    if m:
        return f"{m.group(1).replace('-', ' ').replace('_', ' ').strip()} Permit"

    m = re.match(r"Form_([A-Z0-9\-]+)_(.+?)(?:_\d{8})?$", name)
    if m:
        num = m.group(1).replace("-", " ")
        desc = m.group(2).replace("-", " ").replace("_", " ").strip()
        return f"{desc} Form {num}"

    m = re.match(r"Handbook_(.+?)(?:_\d{8})?$", name)
    if m:
        rest = re.sub(r"^TOW-?", "", m.group(1))
        rest = rest.replace("-", " ").replace("_", " ").strip()
        has_date = bool(re.search(r"_\d{8}$", name))
        return f"{rest} Handbook" + (" (revised)" if has_date else "")

    m = re.match(r"Comm_(.+?)_\d{6,8}$", name)
    if m:
        desc = m.group(1).replace("-", " ").replace("_", " ").strip()
        return f"{desc} Communication"

    return name.replace("_", " ").replace("-", " ").strip()


def parse_date(filename):
    m = re.search(r"(\d{8})", filename)
    if m:
        d = m.group(1)
        return f"{d[:4]}-{d[4:6]}-{d[6:]}"
    # Try YYMMDD format (e.g. 260715 = 2026-07-15)
    m = re.search(r"_(\d{6})(?:\.|$)", filename)
    if m:
        d = m.group(1)
        yy, mm, dd = d[:2], d[2:4], d[4:6]
        if 1 <= int(mm) <= 12 and 1 <= int(dd) <= 31:
            return f"20{yy}-{mm}-{dd}"
    return None


def extract_meeting_slug(filename):
    """Extract a descriptive meeting title from a non-standard filename.
    e.g. Agenda_2027-Initial-Budget-Meeting_20260805.pdf -> '2027 Initial Budget Meeting'
    Returns None if the filename matches standard RTBM/STBM/TBSM patterns."""
    name = filename
    for ext in [".pdf", ".docx", ".xlsx", ".doc", ".xls", ".pptx"]:
        if name.lower().endswith(ext):
            name = name[:-len(ext)]
            break
    # Only applies to Agenda files that aren't standard patterns
    if not name.startswith("Agenda_"):
        return None
    if re.match(r"Agenda_(RTBM|STBM|TBSM|BOR)_", name):
        return None
    if re.match(r"Agenda_\d{8}$", name):
        return None
    # Strip "Agenda_" prefix
    slug = re.sub(r"^Agenda_", "", name)
    # Strip trailing date
    slug = re.sub(r"_?\d{8}$", "", slug)
    # Strip DRAFT suffix
    slug = re.sub(r"_?DRAFT$", "", slug, flags=re.IGNORECASE)
    # Clean separators
    slug = slug.replace("_", " ").replace("-", " ").strip()
    return slug if slug else None


def scan_folder(token, drive_id, folder_path):
    """Scan a meeting folder for documents and subfolders.
    Returns (docs, subfolders) where subfolders is a list of
    {"name": ..., "url": ..., "documents": [...]} dicts.
    Skips any subfolder named 'Internal Only'."""
    try:
        children = list_children(token, drive_id, folder_path)
    except Exception as e:
        print(f"  Warning: could not read {folder_path}: {e}")
        return [], []

    docs = []
    subfolders = []
    folder_url = None
    folder_url_fetched = False

    for item in children:
        name = item["name"]

        # Handle subfolders: skip Internal Only, scan everything else
        if "folder" in item:
            if name.strip().lower() == SKIP_FOLDER_NAME.lower():
                continue
            sub_path = f"{folder_path}/{name}"
            print(f"    Scanning subfolder: {name}...")
            sub_docs = scan_subfolder(token, drive_id, sub_path)
            if sub_docs:
                sub_folder_url = make_folder_link(token, drive_id, sub_path)
                subfolder_entry = {
                    "name": name,
                    "documents": sub_docs,
                }
                if sub_folder_url:
                    subfolder_entry["url"] = sub_folder_url
                subfolders.append(subfolder_entry)
            continue

        try:
            link = make_link(token, drive_id, item["id"])
        except Exception as e:
            print(f"  Warning: skipping {name}: {e}")
            continue

        if not folder_url_fetched:
            folder_url = make_folder_link(token, drive_id, folder_path)
            folder_url_fetched = True

        date = parse_date(name) or item.get("lastModifiedDateTime", "")[:10]
        posted = item.get("createdDateTime", item.get("lastModifiedDateTime", ""))[:10]
        doc = {"label": infer_label(name), "filename": name, "url": link, "date": date, "posted": posted}
        if "DRAFT" in name.upper():
            doc["draft"] = True
        if folder_url:
            doc["folderUrl"] = folder_url
        docs.append(doc)

    def key(d):
        l = d["label"].lower()
        if "agenda" in l: return (0, l)
        if "minutes" in l: return (1, l)
        return (2, l)
    docs.sort(key=key)
    return docs, subfolders


def scan_subfolder(token, drive_id, folder_path):
    """Scan a subfolder inside a meeting folder for files only.
    Returns a list of document dicts. Skips any nested subfolders."""
    try:
        children = list_children(token, drive_id, folder_path)
    except Exception as e:
        print(f"    Warning: could not read subfolder {folder_path}: {e}")
        return []

    docs = []
    for item in children:
        name = item["name"]
        if "folder" in item:
            continue  # Don't recurse deeper inside meeting subfolders

        try:
            link = make_link(token, drive_id, item["id"])
        except Exception as e:
            print(f"    Warning: skipping {name}: {e}")
            continue

        date = parse_date(name) or item.get("lastModifiedDateTime", "")[:10]
        posted = item.get("createdDateTime", item.get("lastModifiedDateTime", ""))[:10]
        doc = {"label": infer_label(name), "filename": name, "url": link, "date": date, "posted": posted}
        if "DRAFT" in name.upper():
            doc["draft"] = True
        docs.append(doc)

    docs.sort(key=lambda d: d["label"].lower())
    return docs


def scan_library(token, drive_id, folder_path=None, folder_label=None, depth=0):
    """Recursively scan a document library (not meeting-folder-shaped).
    Skips any folder named SKIP_FOLDER_NAME at any depth. Recurses into
    other subfolders up to MAX_SCAN_DEPTH levels below the library root.
    Each doc gets a folderUrl pointing at its own parent folder (fetched
    once per folder, not once per file)."""
    try:
        items = list_children(token, drive_id, folder_path) if folder_path else list_root(token, drive_id)
    except Exception as e:
        print(f"  Warning: could not read {folder_path or '(root)'}: {e}")
        return []

    docs = []
    folder_url = None
    folder_url_fetched = False

    for item in items:
        name = item["name"]

        if "folder" in item:
            if name.strip().lower() == SKIP_FOLDER_NAME.lower():
                continue
            if depth < MAX_SCAN_DEPTH:
                sub_path = f"{folder_path}/{name}" if folder_path else name
                docs.extend(scan_library(token, drive_id, sub_path, name, depth + 1))
            continue

        try:
            link = make_link(token, drive_id, item["id"])
        except Exception as e:
            print(f"    Warning: skipping {name}: {e}")
            continue

        if not folder_url_fetched:
            folder_url = make_folder_link(token, drive_id, folder_path)
            folder_url_fetched = True

        base = re.sub(r"\.(pdf|docx?|xlsx?|pptx?)$", "", name, flags=re.IGNORECASE)
        date = parse_date(name) or item.get("lastModifiedDateTime", "")[:10]
        posted = item.get("createdDateTime", item.get("lastModifiedDateTime", ""))[:10]
        doc = {"label": clean_name(base), "filename": name, "url": link, "date": date, "posted": posted}
        if folder_label:
            doc["folder"] = folder_label
        if folder_url:
            doc["folderUrl"] = folder_url
        docs.append(doc)

    return docs


def build_flat_site_data(token, site_config):
    print(f"Locating site: {site_config['key']} ({site_config['path']})...")
    try:
        site_id = get_site_id(token, site_config["path"])
    except Exception as e:
        print(f"  Warning: could not find site {site_config['path']}: {e}")
        return {"libraries": [{"name": lib, "documents": []} for lib in site_config["libraries"]]}

    libraries_out = []
    for lib_name in site_config["libraries"]:
        print(f"  Scanning library: {lib_name}...")
        try:
            drive_id = get_drive_id(token, site_id, lib_name)
        except Exception as e:
            print(f"    Warning: {e}")
            libraries_out.append({"name": lib_name, "documents": []})
            continue
        docs = scan_library(token, drive_id)
        docs.sort(key=lambda d: d["label"].lower())
        print(f"    Found {len(docs)} document(s).")
        libraries_out.append({"name": lib_name, "documents": docs})

    return {"libraries": libraries_out}


def parse_folder_name(name, mtype):
    m = re.search(r"(\d{8})$", name)
    if not m:
        return None
    raw = m.group(1)
    date_str = f"{raw[:4]}-{raw[4:6]}-{raw[6:]}"
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    if mtype == "regular":
        return {"id": f"regular-{date_str}",
                "title": f"Regular Town Board Meeting - {dt.strftime('%B %Y')}",
                "date": date_str}
    else:
        return {"id": f"special-{date_str}",
                "title": f"Special Town Board Meeting - {dt.strftime('%B %-d, %Y')}",
                "date": date_str}


def build_data(token, drive_id):
    meetings = []

    print("Fetching Events Calendar...")
    ical_events = _fetch_ical_events()

    print("Scanning Next Meeting...")
    next_docs, next_subfolders = scan_folder(token, drive_id, NEXT_MEETING_PATH)

    next_date = None
    next_type = "regular"
    next_title = "Upcoming Town Board Meeting"
    next_time = "6:00 PM"
    next_location = "Winchester Town Hall"

    for doc in next_docs:
        fn = doc["filename"]
        m = re.search(r"(\d{8})", fn)
        if m:
            raw = m.group(1)
            next_date = f"{raw[:4]}-{raw[4:6]}-{raw[6:]}"
            dt = datetime.strptime(next_date, "%Y-%m-%d")
            if "RTBM" in fn:
                next_type = "regular"
                next_title = f"Regular Town Board Meeting - {dt.strftime('%B %Y')}"
            elif "STBM" in fn or "TBSM" in fn:
                next_type = "special"
                next_title = f"Special Town Board Meeting - {dt.strftime('%B %-d, %Y')}"
            else:
                # Non-standard meeting — extract descriptive slug from filename
                slug = extract_meeting_slug(fn)
                if slug:
                    next_title = slug
                else:
                    next_title = f"Town Board Meeting - {dt.strftime('%B %-d, %Y')}"
                next_type = "special"
            break

    # Resolve time & location: PDF content is most reliable (directly from
    # the agenda header), iCal is fallback, hardcoded defaults are last resort.
    print("  Extracting meeting details from agenda PDF...")
    pdf_details = extract_details_from_next_meeting(token, drive_id, next_docs)
    if pdf_details.get("time"):
        next_time = pdf_details["time"]
    if pdf_details.get("location"):
        next_location = pdf_details["location"]

    # If PDF didn't yield time, try iCal
    if not pdf_details.get("time"):
        ical_event = get_next_meeting_from_ical(ical_events, match_date=next_date)
        if ical_event:
            next_time = ical_event["time"]
            if not pdf_details.get("location"):
                next_location = ical_event["location"]
            print(f"  iCal fallback: {ical_event['title']} at {ical_event['time']} @ {ical_event['location']}")

    if not next_date:
        next_date = datetime.today().strftime("%Y-%m-%d")

    next_meeting = {
        "id": f"{next_type}-{next_date}", "title": next_title,
        "type": next_type, "status": "upcoming", "date": next_date,
        "time": next_time, "location": next_location,
        "documents": next_docs,
    }
    if next_subfolders:
        next_meeting["subfolders"] = next_subfolders
    meetings.append(next_meeting)

    print("Scanning Previous Regular Meetings...")
    try:
        reg_folders = list_children(token, drive_id, PREV_REGULAR_PATH)
    except Exception as e:
        print(f"  Warning: {e}")
        reg_folders = []
    reg = []
    for item in reg_folders:
        if "folder" not in item:
            continue
        p = parse_folder_name(item["name"], "regular")
        if not p:
            continue
        print(f"  Scanning {item['name']}...")
        docs, subfolders = scan_folder(token, drive_id, f"{PREV_REGULAR_PATH}/{item['name']}")
        entry = {"id": p["id"], "title": p["title"], "type": "regular",
                 "status": "complete", "date": p["date"], "documents": docs}
        if subfolders:
            entry["subfolders"] = subfolders
        reg.append(entry)
    reg.sort(key=lambda x: x["date"], reverse=True)
    meetings.extend(reg)

    print("Scanning Previous Special Meetings...")
    try:
        spec_folders = list_children(token, drive_id, PREV_SPECIAL_PATH)
    except Exception as e:
        print(f"  Warning: {e}")
        spec_folders = []
    spec = []
    for item in spec_folders:
        if "folder" not in item:
            continue
        p = parse_folder_name(item["name"], "special")
        if not p:
            continue
        print(f"  Scanning {item['name']}...")
        docs, subfolders = scan_folder(token, drive_id, f"{PREV_SPECIAL_PATH}/{item['name']}")
        entry = {"id": p["id"], "title": p["title"], "type": "special",
                 "status": "complete", "date": p["date"], "documents": docs}
        if subfolders:
            entry["subfolders"] = subfolders
        spec.append(entry)
    spec.sort(key=lambda x: x["date"], reverse=True)
    meetings.extend(spec)

    return {"meetings": meetings}


def write_github(data):
    headers = {"Authorization": f"Bearer {GH_PAT}", "Accept": "application/vnd.github+json"}
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_FILE}"
    r = requests.get(url, headers=headers)
    sha = r.json().get("sha") if r.status_code == 200 else None
    content = base64.b64encode(json.dumps(data, indent=2, ensure_ascii=False).encode()).decode()
    payload = {
        "message": f"Auto-sync from SharePoint [{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}]",
        "content": content
    }
    if sha:
        payload["sha"] = sha
    r = requests.put(url, headers=headers, json=payload)
    r.raise_for_status()
    print("data.json updated successfully.")


def write_github_file(path, content_str):
    """Write any file to the GitHub repo. Used by permits sync."""
    headers = {"Authorization": f"Bearer {GH_PAT}", "Accept": "application/vnd.github+json"}
    url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{path}"
    r = requests.get(url, headers=headers)
    sha = r.json().get("sha") if r.status_code == 200 else None
    encoded = base64.b64encode(content_str.encode()).decode()
    payload = {
        "message": f"Auto-sync {path} [{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}]",
        "content": encoded
    }
    if sha:
        payload["sha"] = sha
    r = requests.put(url, headers=headers, json=payload)
    r.raise_for_status()
    print(f"{path} updated successfully.")


# ── Permits sync ─────────────────────────────────────────

def _ps(v):
    """Cell value -> clean string."""
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def _pdate(v):
    """Cell value -> YYYY-MM-DD, or '' if unparseable."""
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    t = str(v).strip()
    if not t:
        return ""
    if re.match(r"^\d{4}-\d{2}-\d{2}$", t):
        return t
    m = re.match(r"^(\d{4}-\d{2}-\d{2})T", t)
    if m:
        return m.group(1)
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y"):
        try:
            return datetime.strptime(t, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return ""


def _pfloat(v):
    try:
        return round(float(str(v).strip()), 6)
    except (TypeError, ValueError):
        return None


def _pcoords(v):
    """route_coords cell -> list of [lat, lng] pairs, or None."""
    t = _ps(v)
    if not t:
        return None
    try:
        parsed = json.loads(t)
    except json.JSONDecodeError:
        return None
    if not isinstance(parsed, list) or len(parsed) < 2:
        return None
    out = []
    for pt in parsed:
        if isinstance(pt, (list, tuple)) and len(pt) >= 2:
            lat, lng = _pfloat(pt[0]), _pfloat(pt[1])
            if lat is not None and lng is not None:
                out.append([lat, lng])
    return out if len(out) >= 2 else None


def _plocation(row):
    """Build a public-facing location string from road + address."""
    road = _ps(row.get("road"))
    addr = _ps(row.get("address"))
    if road and addr:
        if road.lower() in addr.lower():
            return addr
        return f"{road} \u2014 {addr}"
    return road or addr or "Location on file with the Town Clerk"


def fetch_permit_rows(token):
    """Download the tracker workbook and return a list of row dicts."""
    url = f"{GRAPH_BASE}/drives/{PERMITS_DRIVE_ID}/items/{PERMITS_ITEM_ID}/content"
    r = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=60)
    r.raise_for_status()

    wb = load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)
    if PERMITS_SHEET not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{PERMITS_SHEET}' not found in workbook")
    ws = wb[PERMITS_SHEET]

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < PERMITS_HEADER_ROW:
        return []

    headers = [_ps(h) for h in rows[PERMITS_HEADER_ROW - 1]]
    out = []
    for raw in rows[PERMITS_FIRST_DATA - 1:]:
        rec = {}
        for i, h in enumerate(headers):
            if h and i < len(raw):
                rec[h] = raw[i]
        if any(_ps(v) for v in rec.values()):
            out.append(rec)
    wb.close()
    return out


def build_permits_data(rows):
    """Convert tracker rows into the public permits.json structure.
    Only published fields are included — no applicant PII."""
    permits = []
    skipped = {"status": 0, "geo": 0, "dates": 0}

    for row in rows:
        status = _ps(row.get("map_status")).lower()

        if _ps(row.get("permit_id")).upper().startswith("EXAMPLE"):
            continue

        if status not in PERMITS_PUBLISHED:
            skipped["status"] += 1
            continue

        start = _pdate(row.get("authorized_start")) or _pdate(row.get("start_date"))
        end   = _pdate(row.get("authorized_end"))   or _pdate(row.get("end_date"))
        if not start or not end:
            skipped["dates"] += 1
            continue
        if end < start:
            start, end = end, start

        pid_for_url = row.get("_pdf_id") or _ps(row.get("permit_id")) or f"P{len(permits)+1:04d}"
        entry = {
            "id":           pid_for_url,
            "permitNumber": _ps(row.get("permit_number")),
            "title":        _ps(row.get("title")) or _ps(row.get("type")) or "Permitted work",
            "org":          _ps(row.get("org")) or "Applicant on file",
            "location":     _plocation(row),
            "startDate":    start,
            "endDate":      end,
            "status":       status,
            "jurisdiction": "town",
            "pdfUrl":       f"/{PDF_PUBLIC_PATH}/{pid_for_url}.pdf",
        }

        traffic = _ps(row.get("traffic"))
        if traffic:
            entry["traffic"] = traffic

        cname = _ps(row.get("public_contact_name"))
        cphone = _ps(row.get("public_contact_phone"))
        if cname:
            entry["contactName"] = cname
        if cphone:
            entry["contactPhone"] = cphone

        geo_type = _ps(row.get("geo_type")).lower()
        coords = _pcoords(row.get("route_coords"))
        lat, lng = _pfloat(row.get("lat")), _pfloat(row.get("lng"))

        if geo_type == "line" and coords:
            entry["geoType"] = "line"
            entry["coordinates"] = coords
            mid = coords[len(coords) // 2]
            entry["lat"], entry["lng"] = mid[0], mid[1]
        elif lat is not None and lng is not None:
            entry["geoType"] = "point"
            entry["lat"], entry["lng"] = lat, lng
        elif coords:
            entry["geoType"] = "line"
            entry["coordinates"] = coords
            mid = coords[len(coords) // 2]
            entry["lat"], entry["lng"] = mid[0], mid[1]
        else:
            skipped["geo"] += 1
            continue

        permits.append(entry)

    order = {"proposed": 1, "approved": 0}
    permits.sort(key=lambda p: (order.get(p["status"], 2), p["startDate"]))

    return {
        "updated": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "permits": permits,
    }, skipped


def sync_permits(token):
    """Fetch permit tracker from SharePoint, build permits.json, push to GitHub."""
    print("\n── Permits ──")
    rows = fetch_permit_rows(token)
    print(f"   {len(rows)} row(s) in tracker")

    # Assign one stable ID per row up front so JSON, PDFs, and state files all agree.
    for i, row in enumerate(rows, 1):
        row["_pdf_id"] = _ps(row.get("permit_id")) or f"P{i:04d}"
        row["__row"] = i + 2   # sheet row (header is row 2, data starts at row 3)

    data, skipped = build_permits_data(rows)
    print(f"   {len(data['permits'])} published")
    if skipped["status"]:
        print(f"   {skipped['status']} withheld (status not proposed/approved)")
    if skipped["dates"]:
        print(f"   {skipped['dates']} skipped — missing or invalid dates")
    if skipped["geo"]:
        print(f"   {skipped['geo']} skipped — no map location")

    payload = json.dumps(data, indent=2, ensure_ascii=False)
    write_github_file(PERMITS_GITHUB_FILE, payload)

    print("   Checking for status-change notifications...")
    try:
        sent, notify_skipped = notify_status_changes(token, rows)
        if sent:
            print(f"   {sent} email(s) sent")
        else:
            print("   No status changes to notify")
        if notify_skipped:
            print(f"   {notify_skipped} skipped (no applicant email)")

        # Commit the updated state file so next run remembers what we notified.
        import os as _os
        if _os.path.exists("permit_notify_state.json"):
            with open("permit_notify_state.json", "r", encoding="utf-8") as _f:
                write_github_file("permit_notify_state.json", _f.read())
    except Exception as e:
        print(f"   ! notification step failed: {e}")
        print("   (permits.json was still published)")

    print("   Generating permit record PDFs...")
    try:
        pub_pdfs, int_pdfs, pdf_unchanged, pdf_errors = generate_permit_pdfs(rows, token=token)
        if pub_pdfs or int_pdfs:
            print(f"   {pub_pdfs} public + {int_pdfs} internal PDF(s) rebuilt")
        else:
            print("   No PDFs needed rebuilding (content unchanged)")
        if pdf_errors:
            print(f"   {pdf_errors} PDF error(s) — see above")

        # Commit the PDF state file so next run knows what content hashes it saw.
        import os as _os2
        if _os2.path.exists("permit_pdf_state.json"):
            with open("permit_pdf_state.json", "r", encoding="utf-8") as _f2:
                write_github_file("permit_pdf_state.json", _f2.read())
    except Exception as e:
        print(f"   ! PDF step failed: {e}")
        print("   (permits.json was still published)")


def main():
    print("Authenticating with Microsoft Graph...")
    token = get_token()
    print("Locating SharePoint site and drive...")
    site_id = get_site_id(token)
    drive_id = get_drive_id(token, site_id)
    print("Building data.json from SharePoint...")
    data = build_data(token, drive_id)
    total = sum(len(m["documents"]) for m in data["meetings"])
    sub_total = sum(
        sum(len(sf["documents"]) for sf in m.get("subfolders", []))
        for m in data["meetings"]
    )
    print(f"Found {len(data['meetings'])} meetings, {total} documents + {sub_total} subfolder documents.")

    print("Scanning additional sites...")
    sites_data = {}
    for site_config in FLAT_SITES:
        sites_data[site_config["key"]] = build_flat_site_data(token, site_config)
    data["sites"] = sites_data
    sites_total = sum(
        len(lib["documents"])
        for site in sites_data.values()
        for lib in site["libraries"]
    )
    print(f"Found {sites_total} document(s) across {len(FLAT_SITES)} additional site(s).")

    print("Writing to GitHub...")
    write_github(data)

    sync_permits(token)

    print("Done.")


if __name__ == "__main__":
    main()
