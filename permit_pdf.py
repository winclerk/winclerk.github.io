"""
permit_pdf.py — generate permit record PDFs.

Produces TWO PDFs per permit:
  * PUBLIC  — omits applicant PII (phone, email, parcel, signature, internal notes),
              published to the winclerk.github.io repo as public/permits/<id>.pdf
              and linked from the permit card on winchesterwi.com/permits.
  * INTERNAL — full record including PII, uploaded to Records Archive → Permits →
               "Permit Records" folder in SharePoint. For Board packets and clerk use.

Content-change detection: hashes the field set that each PDF renders; only
regenerates when the hash differs from what's stored in permit_pdf_state.json.

Called from sync.py after the tracker rows are fetched. Can also be run
standalone via PERMIT=P0001 python permit_pdf.py (single-permit mode) or
PERMIT=ALL to rebuild everything.
"""

import hashlib
import io
import json
import os
import re
import unicodedata
from datetime import datetime, date, timezone

import requests

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    BaseDocTemplate, Frame, PageTemplate, Paragraph, Spacer,
    Table, TableStyle, Image,
)

# ── Config ───────────────────────────────────────────────────────────
TENANT_ID     = os.environ.get("AZURE_TENANT_ID", "")
CLIENT_ID     = os.environ.get("AZURE_CLIENT_ID", "")
CLIENT_SECRET = os.environ.get("AZURE_CLIENT_SECRET", "")

# GitHub — for publishing public PDFs to winclerk.github.io
GITHUB_TOKEN  = os.environ.get("GH_PAT", "") or os.environ.get("GITHUB_TOKEN", "")
GITHUB_REPO   = os.environ.get("GITHUB_REPO", "winclerk/winclerk.github.io")
GITHUB_BRANCH = os.environ.get("GITHUB_BRANCH", "main")
PUBLIC_PATH   = "public/permits"  # rendered URL: winclerk.github.io/public/permits/<id>.pdf

# SharePoint — internal PDFs land in Records Archive → Permits → "Permit Records"
PERMITS_DRIVE_ID = "b!c95LT9gy6kqiItDGFto7RFnHF7mxITJGsCZJt01CCvi1TQvoZoFtT7YMKBgrUG67"
PERMITS_ITEM_ID  = "01UVO6XCSTSHOVIGOYBJE3ZIQJYCZ3YSTV"
INTERNAL_FOLDER  = "Permit Records"

LOGO_PATH  = "assets/winchester-logo.png"   # optional
STATE_FILE = "permit_pdf_state.json"

GRAPH_BASE = "https://graph.microsoft.com/v1.0"

# Winchester palette
TEAL_DARK = colors.HexColor("#193C3C")
TEAL_MID  = colors.HexColor("#00505A")
TAN       = colors.HexColor("#F5F2EC")
GOLD      = colors.HexColor("#A59664")
GREY      = colors.HexColor("#666666")
RULE      = colors.HexColor("#D8D8D8")


# ── Field layouts — one for public, one for internal ────────────────

_APPLICATION = ("Application Record", [
    ("permit_id",        "Permit ID"),
    ("permit_number",    "Permit number"),
    ("submitted",        "Submitted"),
    ("map_status",       "Status"),
    ("board_date",       "Board meeting date"),
    ("authorized_start", "Authorized start"),
    ("authorized_end",   "Authorized end"),
])

_PROJECT = ("Project", [
    ("type",        "Permit type"),
    ("title",       "Public title"),
    ("description", "Scope of work"),
    ("traffic",     "Traffic impact"),
])

_LOCATION_PUBLIC = ("Location", [
    ("road",     "Road"),
    ("address",  "Address"),
    ("geo_type", "Geometry"),
    ("lat",      "Latitude"),
    ("lng",      "Longitude"),
])

_LOCATION_INTERNAL = ("Location", [
    ("road",     "Road"),
    ("address",  "Address"),
    ("parcel",   "Parcel number"),
    ("geo_type", "Geometry"),
    ("lat",      "Latitude"),
    ("lng",      "Longitude"),
])

_SCHEDULE = ("Schedule", [
    ("start_date", "Requested start"),
    ("end_date",   "Requested end"),
])

_CONTACTS_PUBLIC = ("Public Contact", [
    ("org",                  "Organization"),
    ("contractor",           "Contractor"),
    ("public_contact_name",  "On-site contact"),
    ("public_contact_phone", "On-site contact phone"),
])

_CONTACTS_INTERNAL = ("Applicant and Contacts", [
    ("org",                  "Organization"),
    ("applicant",            "Applicant"),
    ("phone",                "Phone"),
    ("email",                "Email"),
    ("contractor",           "Contractor"),
    ("property_owner",       "Property owner"),
    ("public_contact_name",  "Public contact"),
    ("public_contact_phone", "Public contact phone"),
])

_CONSTRUCTION = ("Construction Details", [
    ("method",                "Installation method"),
    ("depth",                 "Depth"),
    ("crossings",             "Road crossings"),
    ("surface",               "Surface"),
    ("width",                 "Width"),
    ("lotline_distance",      "Distance to lot line"),
    ("slope",                 "Slope"),
    ("culvert_exists",        "Existing culvert"),
    ("culvert_existing_size", "Existing culvert size"),
])

_COMPLIANCE_PUBLIC = ("Compliance", [
    ("insurance",      "Certificate of insurance filed"),
    ("diggers_ticket", "Diggers Hotline ticket"),
])

_COMPLIANCE_INTERNAL = ("Compliance", [
    ("insurance",      "Certificate of insurance"),
    ("diggers_ticket", "Diggers Hotline ticket"),
])

_CERTIFICATION_INTERNAL = ("Applicant Certification", [
    ("signature",      "Signed by"),
    ("signature_date", "Date signed"),
])

_DECISION_PUBLIC = ("Board Decision & Conditions", [
    ("conditions", "Conditions of approval"),
])

_CLERK_INTERNAL = ("Clerk Use Only", [
    ("fee_paid",    "Fee paid"),
    ("conditions",  "Conditions of approval"),
    ("clerk_notes", "Clerk notes"),
])


PUBLIC_SECTIONS = [
    _APPLICATION,
    _PROJECT,
    _LOCATION_PUBLIC,
    _SCHEDULE,
    _CONTACTS_PUBLIC,
    _CONSTRUCTION,
    _COMPLIANCE_PUBLIC,
    _DECISION_PUBLIC,
]

INTERNAL_SECTIONS = [
    _APPLICATION,
    _PROJECT,
    _LOCATION_INTERNAL,
    _SCHEDULE,
    _CONTACTS_INTERNAL,
    _CONSTRUCTION,
    _COMPLIANCE_INTERNAL,
    _CERTIFICATION_INTERNAL,
    _CLERK_INTERNAL,
]

# Fields whose columns feed the content-change hash.
_HASHED_COLUMNS = [
    "permit_id", "permit_number", "submitted", "map_status",
    "board_date", "authorized_start", "authorized_end",
    "conditions", "clerk_notes", "fee_paid",
    "type", "title", "description", "traffic",
    "road", "address", "parcel", "geo_type", "lat", "lng", "route_coords",
    "start_date", "end_date",
    "org", "applicant", "phone", "email", "contractor", "property_owner",
    "public_contact_name", "public_contact_phone",
    "method", "depth", "crossings", "surface", "width",
    "lotline_distance", "slope", "culvert_exists", "culvert_existing_size",
    "insurance", "diggers_ticket", "signature", "signature_date",
]

_PRIVATE_ONLY_COLUMNS = {"applicant", "phone", "email", "parcel",
                         "signature", "signature_date", "clerk_notes", "fee_paid"}


# ── Value formatting ─────────────────────────────────────────────────
def _s(v):
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def _pretty(col, v):
    t = _s(v)
    if not t:
        return "\u2014"

    if col == "submitted":
        m = re.match(r"^(\d{4})-(\d{2})-(\d{2})T(\d{2}):(\d{2})", t)
        if m:
            y, mo, d, hh, mm = m.groups()
            try:
                dt = datetime(int(y), int(mo), int(d), int(hh), int(mm), tzinfo=timezone.utc)
                return f"{dt.strftime('%B')} {dt.day}, {dt.year} at {dt.strftime('%H:%M')} UTC"
            except ValueError:
                return t
    if col in ("map_status", "geo_type"):
        return t.capitalize()
    if col in ("insurance", "diggers_ticket", "culvert_exists", "fee_paid"):
        low = t.lower()
        if low in ("1", "true", "yes", "y"):
            return "Yes"
        if low in ("0", "false", "no", "n"):
            return "No"
    if col in ("lat", "lng"):
        try:
            return f"{float(t):.6f}"
        except ValueError:
            return t
    return t


def _clean_filename(s):
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^A-Za-z0-9 _.-]+", "", s).strip()
    s = re.sub(r"\s+", "_", s)
    return s[:80] or "Permit"


def _content_hash(row, variant):
    payload = {"variant": variant}
    for col in _HASHED_COLUMNS:
        if variant == "public" and col in _PRIVATE_ONLY_COLUMNS:
            continue
        payload[col] = _s(row.get(col))
    blob = json.dumps(payload, sort_keys=True, ensure_ascii=False).encode("utf-8")
    return hashlib.sha256(blob).hexdigest()


# ── PDF rendering ────────────────────────────────────────────────────
BODY = ParagraphStyle("body", fontName="Helvetica", fontSize=9, leading=12,
                      textColor=colors.HexColor("#222222"))
LABEL = ParagraphStyle("label", fontName="Helvetica-Bold", fontSize=8,
                       leading=11, textColor=GREY)
SECTION = ParagraphStyle("section", fontName="Helvetica-Bold", fontSize=10.5,
                         leading=14, textColor=colors.white)
TITLE = ParagraphStyle("title", fontName="Helvetica-Bold", fontSize=17,
                       leading=21, textColor=TEAL_DARK)
SUB = ParagraphStyle("sub", fontName="Helvetica", fontSize=9.5, leading=13,
                     textColor=GREY)
EYEBROW = ParagraphStyle("eyebrow", fontName="Helvetica-Bold", fontSize=7.5,
                         leading=10, textColor=GOLD)
FOOT = ParagraphStyle("foot", fontName="Helvetica", fontSize=7.5, leading=10,
                      textColor=GREY, alignment=TA_CENTER)


def _section_block(title, pairs, row, accent=TEAL_MID):
    live = [(lbl, _pretty(col, row.get(col))) for col, lbl in pairs]
    if all(v == "\u2014" for _, v in live):
        return None

    data = [[Paragraph(title.upper(), SECTION), ""]]
    data += [[Paragraph(lbl, LABEL), Paragraph(val.replace("\n", "<br/>"), BODY)]
             for lbl, val in live]

    tbl = Table(data, colWidths=[2.0 * inch, 4.9 * inch], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("SPAN",          (0, 0), (-1, 0)),
        ("BACKGROUND",    (0, 0), (-1, 0), accent),
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
        ("TOPPADDING",    (0, 0), (-1, 0), 5),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 5),
        ("TOPPADDING",    (0, 1), (-1, -1), 3.5),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 3.5),
        ("LINEBELOW",     (0, 1), (-1, -2), 0.4, RULE),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, TAN]),
    ]))
    return [tbl, Spacer(1, 9)]


def _route_block(row):
    coords_raw = _s(row.get("route_coords"))
    if not coords_raw:
        return None
    try:
        pts = json.loads(coords_raw)
        if not isinstance(pts, list) or not pts:
            return None
        lines = [f"{i + 1}.&nbsp;&nbsp;{float(p[0]):.6f}, {float(p[1]):.6f}"
                 for i, p in enumerate(pts) if isinstance(p, (list, tuple)) and len(p) >= 2]
    except Exception:
        return None
    if not lines:
        return None

    data = [[Paragraph("ROUTE VERTICES", SECTION)],
            [Paragraph("&nbsp;&nbsp;&nbsp;".join(lines), BODY)]]
    tbl = Table(data, colWidths=[6.9 * inch], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), TEAL_MID),
        ("BACKGROUND",    (0, 1), (-1, -1), colors.white),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    return [tbl, Spacer(1, 9)]


def _header_footer_public(canvas, doc):
    _header_footer(canvas, doc,
                   subheading="Permit Record \u2014 Public",
                   footer="This is a public summary. Applicant contact details are omitted; "
                          "request the internal record from the Town Clerk under Wis. Stat. ch. 19 if needed.")


def _header_footer_internal(canvas, doc):
    _header_footer(canvas, doc,
                   subheading="Permit Record \u2014 Internal",
                   footer="This internal record contains applicant contact information and signature. "
                          "Review under Wis. Stat. ch. 19 before public release.")


def _header_footer(canvas, doc, subheading, footer):
    canvas.saveState()
    w, h = LETTER
    canvas.setFillColor(TEAL_DARK)
    canvas.rect(0, h - 0.42 * inch, w, 0.42 * inch, stroke=0, fill=1)
    canvas.setFillColor(GOLD)
    canvas.rect(0, h - 0.47 * inch, w, 0.05 * inch, stroke=0, fill=1)

    canvas.setFillColor(colors.white)
    canvas.setFont("Helvetica-Bold", 9)
    canvas.drawString(0.75 * inch, h - 0.29 * inch, "TOWN OF WINCHESTER, WISCONSIN")
    canvas.setFont("Helvetica", 8)
    canvas.drawRightString(w - 0.75 * inch, h - 0.29 * inch, subheading)

    canvas.setFillColor(GREY)
    canvas.setFont("Helvetica", 7.5)
    canvas.drawCentredString(w / 2, 0.58 * inch, footer)
    canvas.drawCentredString(w / 2, 0.42 * inch,
        "Town of Winchester  \u00b7  7228 CTH W, Winchester, WI 54557  \u00b7  "
        "715-686-2123  \u00b7  clerk@winchester.wi.gov")
    canvas.drawCentredString(w / 2, 0.27 * inch, f"Page {doc.page}")
    canvas.restoreState()


def _build(row, sections, variant, header_footer_fn, subtitle_suffix=""):
    buf = io.BytesIO()
    doc = BaseDocTemplate(
        buf, pagesize=LETTER,
        leftMargin=0.75 * inch, rightMargin=0.75 * inch,
        topMargin=0.85 * inch, bottomMargin=0.85 * inch,
        title=f"Permit Record {_s(row.get('permit_id')) or row.get('__row')}",
        author="Town of Winchester",
    )
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id="f")
    doc.addPageTemplates([PageTemplate(id="main", frames=[frame], onPage=header_footer_fn)])

    story = []
    if os.path.exists(LOGO_PATH):
        try:
            logo = Image(LOGO_PATH, width=0.85 * inch, height=0.85 * inch)
            logo.hAlign = "LEFT"
            story += [logo, Spacer(1, 6)]
        except Exception:
            pass

    kind = _s(row.get("type")) or "Permit"
    story.append(Paragraph(kind.upper(), EYEBROW))
    story.append(Paragraph(_s(row.get("title")) or "Permit Application", TITLE))

    where = " \u2014 ".join([x for x in [_s(row.get("road")), _s(row.get("address"))] if x])
    ident = _s(row.get("permit_number")) or _s(row.get("permit_id")) or f"Sheet row {row.get('__row')}"
    sub = f"{where or 'Location on file'}<br/>Record: {ident}"
    if subtitle_suffix:
        sub += f"<br/><i>{subtitle_suffix}</i>"
    story.append(Paragraph(sub, SUB))
    story.append(Spacer(1, 12))

    for title, pairs in sections:
        blk = _section_block(title, pairs, row)
        if blk:
            story.extend(blk)
        if title == "Location":
            rb = _route_block(row)
            if rb:
                story.extend(rb)

    stamp = datetime.now(timezone.utc).strftime("%B %d, %Y at %H:%M UTC")
    variant_note = "Public summary" if variant == "public" else "Complete internal record"
    story.append(Spacer(1, 6))
    story.append(Paragraph(
        f"{variant_note}. Generated {stamp} from the Town of Winchester permit tracker.",
        FOOT))

    doc.build(story)
    return buf.getvalue()


def build_public_pdf(row):
    return _build(row, PUBLIC_SECTIONS, "public", _header_footer_public,
                  subtitle_suffix="Public summary \u2014 some applicant details omitted")


def build_internal_pdf(row):
    return _build(row, INTERNAL_SECTIONS, "internal", _header_footer_internal)


# ── Graph helpers (internal PDF → SharePoint) ────────────────────────
def _get_token():
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    r = requests.post(url, data={
        "client_id":     CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "grant_type":    "client_credentials",
        "scope":         "https://graph.microsoft.com/.default",
    }, timeout=60)
    r.raise_for_status()
    return r.json()["access_token"]


def _ensure_folder(token, name):
    hdrs = {"Authorization": f"Bearer {token}"}
    r = requests.get(f"{GRAPH_BASE}/drives/{PERMITS_DRIVE_ID}/root/children",
                     headers=hdrs, timeout=60)
    r.raise_for_status()
    for item in r.json().get("value", []):
        if item.get("name") == name and "folder" in item:
            return item["id"]
    r = requests.post(
        f"{GRAPH_BASE}/drives/{PERMITS_DRIVE_ID}/root/children",
        headers={**hdrs, "Content-Type": "application/json"},
        json={"name": name, "folder": {}, "@microsoft.graph.conflictBehavior": "return"},
        timeout=60,
    )
    r.raise_for_status()
    return r.json()["id"]


def _upload_to_sharepoint(token, folder_id, filename, data):
    url = (f"{GRAPH_BASE}/drives/{PERMITS_DRIVE_ID}/items/{folder_id}:/"
           f"{requests.utils.quote(filename)}:/content")
    r = requests.put(url, headers={
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/pdf",
    }, data=data, timeout=180)
    r.raise_for_status()
    return r.json().get("webUrl", "")


# ── GitHub helpers (public PDF → winclerk.github.io) ────────────────
def _github_put_binary(path, data, message):
    import base64 as _b64
    api = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{path}"
    hdrs = {
        "Authorization": f"Bearer {GITHUB_TOKEN}",
        "Accept": "application/vnd.github+json",
    }
    sha = None
    r = requests.get(api, headers=hdrs, params={"ref": GITHUB_BRANCH}, timeout=60)
    if r.status_code == 200:
        sha = r.json().get("sha")
    body = {
        "message": message,
        "content": _b64.b64encode(data).decode("ascii"),
        "branch":  GITHUB_BRANCH,
    }
    if sha:
        body["sha"] = sha
    r = requests.put(api, headers=hdrs, json=body, timeout=120)
    r.raise_for_status()
    return f"https://raw.githubusercontent.com/{GITHUB_REPO}/{GITHUB_BRANCH}/{path}"


# ── State (per-permit content hashes) ────────────────────────────────
def _load_state():
    if not os.path.exists(STATE_FILE):
        return {}
    try:
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f).get("permits", {})
    except Exception:
        return {}


def _save_state(state):
    payload = {
        "updated": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "permits": state,
    }
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)


def _permit_key(row):
    pid = _s(row.get("permit_number")) or _s(row.get("_pdf_id")) or _s(row.get("permit_id"))
    if pid:
        return pid
    sub = _s(row.get("submitted"))
    return f"sub:{sub}" if sub else ""


# ── Entry point called from sync.py ──────────────────────────────────
def generate_for_rows(rows, token=None, force=False):
    """Generates public + internal PDFs for each row whose content has changed.
       Returns (public_built, internal_built, skipped_unchanged, skipped_error)."""
    if token is None:
        token = _get_token()

    state = _load_state()
    new_state = dict(state)
    folder_id = None

    pub_built = 0
    int_built = 0
    skipped = 0
    errors = 0

    for row in rows:
        key = _permit_key(row)
        if not key:
            skipped += 1
            continue

        pub_hash = _content_hash(row, "public")
        int_hash = _content_hash(row, "internal")
        prev = state.get(key, {})

        need_pub = force or prev.get("public") != pub_hash
        need_int = force or prev.get("internal") != int_hash

        if not need_pub and not need_int:
            skipped += 1
            continue

        ident_pub = (_s(row.get("permit_number"))
                     or _s(row.get("_pdf_id"))
                     or _s(row.get("permit_id"))
                     or f"row{row.get('__row', '')}")
        road = _s(row.get("road")) or _s(row.get("title")) or "Permit"
        base = _clean_filename(f"Permit {ident_pub} {road}")

        row_state = dict(prev)

        if need_pub:
            try:
                pdf = build_public_pdf(row)
                public_name = f"{_clean_filename(ident_pub)}.pdf"
                path = f"{PUBLIC_PATH}/{public_name}"
                _github_put_binary(path, pdf,
                                   f"Publish public permit record {ident_pub}")
                print(f"     public   \u2192 {path}")
                pub_built += 1
                row_state["public"] = pub_hash
            except Exception as e:
                print(f"     ! public failed for {key}: {e}")
                errors += 1

        if need_int:
            try:
                pdf = build_internal_pdf(row)
                if folder_id is None:
                    folder_id = _ensure_folder(token, INTERNAL_FOLDER)
                internal_name = f"{base}.pdf"
                _upload_to_sharepoint(token, folder_id, internal_name, pdf)
                print(f"     internal \u2192 {INTERNAL_FOLDER}/{internal_name}")
                int_built += 1
                row_state["internal"] = int_hash
            except Exception as e:
                print(f"     ! internal failed for {key}: {e}")
                errors += 1

        new_state[key] = row_state

    _save_state(new_state)
    return pub_built, int_built, skipped, errors


# ── Standalone mode (env PERMIT=... python permit_pdf.py) ───────────
def _select_rows(rows, selector):
    sel = (selector or "").strip()
    if not sel or sel.upper() == "ALL":
        return rows
    m = re.match(r"^row:(\d+)$", sel, re.I)
    if m:
        want = int(m.group(1))
        return [r for r in rows if r.get("__row") == want]
    low = sel.lower()
    hits = [r for r in rows if _s(r.get("permit_id")).lower() == low]
    if hits: return hits
    hits = [r for r in rows if _s(r.get("permit_number")).lower() == low]
    if hits: return hits
    return [r for r in rows if low in _s(r.get("road")).lower()
            or low in _s(r.get("applicant")).lower()]


def _fetch_rows_standalone(token):
    from openpyxl import load_workbook
    url = f"{GRAPH_BASE}/drives/{PERMITS_DRIVE_ID}/items/{PERMITS_ITEM_ID}/content"
    r = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=120)
    r.raise_for_status()
    wb = load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)
    ws = wb["Permits"]
    raw = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(raw) < 3:
        return []
    headers = [_s(h) for h in raw[1]]
    out = []
    for offset, line in enumerate(raw[2:]):
        rec = {}
        for i, h in enumerate(headers):
            if h and i < len(line):
                rec[h] = line[i]
        if any(_s(v) for v in rec.values()):
            rec["__row"] = 3 + offset
            out.append(rec)
    return out


def main():
    selector = os.environ.get("PERMIT", "ALL")
    force    = os.environ.get("FORCE", "").lower() in ("1", "true", "yes")
    print(f"Selector: {selector}  Force: {force}")

    token = _get_token()
    rows = _fetch_rows_standalone(token)
    print(f"{len(rows)} row(s) in tracker")

    picked = _select_rows(rows, selector)
    if not picked:
        raise SystemExit(f"No permit matched '{selector}'")
    print(f"{len(picked)} row(s) selected")

    pub, intl, skipped, errors = generate_for_rows(picked, token=token, force=force)
    print(f"Public rebuilt: {pub}   Internal rebuilt: {intl}   Unchanged: {skipped}   Errors: {errors}")


if __name__ == "__main__":
    main()
