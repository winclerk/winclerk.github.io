"""
Winchester Permit Sync
──────────────────────
Reads the three permit tracker workbooks from SharePoint (Records Archive → Permits),
auto-assigns permit numbers to any new (blank-permit_number) rows, and writes
permits.json to the winclerk.github.io repo.

  ROW  ->  Right-of-Way.xlsx          (sheet: ROW)
  DW   ->  Driveway.xlsx              (sheet: DW)
  RC   ->  Road Construction.xlsx     (sheet: RC)

Called from sync.py. Reuses get_token() and a write_github_file(path, content)
helper from the existing sync pipeline.

Permit numbering: YYYY-TYPE-N, sequence per type per year, resets January 1.
  Examples: 2026-ROW-1, 2026-DW-3, 2026-RC-2

Privacy — Driveway permits are treated more carefully than ROW/RC:
  - No applicant name, no parcel address in the public payload
  - Location shown as road-level only ("New driveway installation on <road>")
  - Full record still available on request under Wis. Stat. ch. 19

Requires: openpyxl (already in the GitHub Actions workflow)
"""

import io
import json
import re
from datetime import datetime, date

import requests
from openpyxl import load_workbook

# ══════════════════════════════════════════════════════════════
# CONFIG — three trackers, all on the same drive
# ══════════════════════════════════════════════════════════════
PERMITS_DRIVE_ID = "b!c95LT9gy6kqiItDGFto7RFnHF7mxITJGsCZJt01CCvi1TQvoZoFtT7YMKBgrUG67"

PERMIT_SOURCES = [
    {
        "type": "row",
        "label": "Right-of-Way",
        "item_id": "01UVO6XCSTSHOVIGOYBJE3ZIQJYCZ3YSTV",
        "sheet_hint": "ROW",
        "type_code": "ROW",
    },
    {
        "type": "driveway",
        "label": "Driveway",
        "item_id": "01UVO6XCXZ4OT5KA376BBJUTCPENRSMERP",
        "sheet_hint": "DW",
        "type_code": "DW",
    },
    {
        "type": "road_construction",
        "label": "Road Construction",
        "item_id": "01UVO6XCWGHFSBESJDKVGZZBHO2XHQBZDR",
        "sheet_hint": "RC",
        "type_code": "RC",
    },
]

PUBLISHED_STATUSES = {"proposed", "approved"}
EXAMPLE_TOKENS = ("EXAMPLE", "TEMPLATE", "SAMPLE")
GRAPH = "https://graph.microsoft.com/v1.0"


# ══════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════
def _s(v):
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def _date(v):
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    t = str(v).strip()
    if not t:
        return ""
    if re.match(r"^\d{4}-\d{2}-\d{2}$", t):
        return t
    m = re.match(r"^(\d{4}-\d{2}-\d{2})T", t)
    if m:
        return m.group(1)
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y"):
        try:
            return datetime.strptime(t, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return ""


def _float(v):
    try:
        return round(float(str(v).strip()), 6)
    except (TypeError, ValueError):
        return None


def _coords(v):
    t = _s(v)
    if not t:
        return None
    try:
        parsed = json.loads(t)
    except json.JSONDecodeError:
        return None
    if not isinstance(parsed, list) or len(parsed) < 1:
        return None
    out = []
    for pt in parsed:
        if isinstance(pt, (list, tuple)) and len(pt) >= 2:
            lat, lng = _float(pt[0]), _float(pt[1])
            if lat is not None and lng is not None:
                out.append([lat, lng])
    return out or None


def _col_letter(n):
    """1-indexed column number to Excel letter. 1 -> A, 2 -> B, 27 -> AA."""
    letters = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


# ══════════════════════════════════════════════════════════════
# FETCH WORKBOOK — auto-detect header row
# ══════════════════════════════════════════════════════════════
_HEADER_TOKENS = {
    "permit_number", "submitted", "map_status", "board_date",
    "type", "title", "road", "applicant", "geo_type",
}


def _detect_header_row(rows, max_scan=5):
    best_row = 1
    best_score = 0
    for i in range(min(max_scan, len(rows))):
        score = sum(1 for cell in rows[i] if _s(cell).lower() in _HEADER_TOKENS)
        if score > best_score:
            best_score = score
            best_row = i + 1
    return best_row if best_score > 0 else None


def _pick_sheet(wb, sheet_hint):
    if sheet_hint:
        for name in wb.sheetnames:
            if name.lower() == sheet_hint.lower():
                return name
    return wb.sheetnames[0] if wb.sheetnames else None


def fetch_source(token, source):
    """Download one workbook and return (rows_as_dicts, meta_dict).
       rows_as_dicts includes a '_excel_row' key on each row for PATCH targeting.
       meta_dict has: header_row, sheet_name, permit_num_col_idx.
       Returns (None, None) on error."""
    url = f"{GRAPH}/drives/{PERMITS_DRIVE_ID}/items/{source['item_id']}/content"
    try:
        r = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=60)
        r.raise_for_status()
    except Exception as e:
        print(f"   ! {source['label']}: could not download workbook ({e})")
        return None, None

    try:
        wb = load_workbook(io.BytesIO(r.content), data_only=True, read_only=True)
    except Exception as e:
        print(f"   ! {source['label']}: could not open workbook ({e})")
        return None, None

    sheet_name = _pick_sheet(wb, source.get("sheet_hint"))
    if not sheet_name:
        print(f"   ! {source['label']}: no sheets in workbook")
        wb.close()
        return None, None
    if source.get("sheet_hint") and sheet_name.lower() != source["sheet_hint"].lower():
        print(f"   ! {source['label']}: sheet '{source['sheet_hint']}' not found, using '{sheet_name}'")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return [], {"header_row": None, "sheet_name": sheet_name, "permit_num_col_idx": None}

    header_row = _detect_header_row(rows)
    if header_row is None:
        print(f"   ! {source['label']}: could not identify a header row")
        wb.close()
        return None, None

    headers = [_s(h) for h in rows[header_row - 1]]
    out = []
    for row_i, raw in enumerate(rows[header_row:], start=header_row + 1):
        rec = {"_excel_row": row_i}
        for i, h in enumerate(headers):
            if h and i < len(raw):
                rec[h] = raw[i]
        if any(_s(v) for k, v in rec.items() if k != "_excel_row"):
            out.append(rec)

    permit_num_col_idx = None
    for i, h in enumerate(headers):
        if h.lower() == "permit_number":
            permit_num_col_idx = i + 1
            break

    meta = {"header_row": header_row, "sheet_name": sheet_name, "permit_num_col_idx": permit_num_col_idx}
    wb.close()
    return out, meta


# ══════════════════════════════════════════════════════════════
# ASSIGN permit_number to blank rows
# ══════════════════════════════════════════════════════════════
def assign_permit_numbers(token, source, rows, meta):
    """Fill in permit_number on any row where it's blank.
       Format: YYYY-TYPE-N (e.g. 2026-ROW-3).
       Writes updates back to SharePoint via Graph API cell PATCH.
       Mutates in-memory rows so downstream code sees new numbers immediately.
       Returns count of numbers assigned."""
    if not rows or not meta or not meta.get("permit_num_col_idx") or not meta.get("sheet_name"):
        return 0

    type_code = source["type_code"]
    year = datetime.utcnow().strftime("%Y")
    prefix = f"{year}-{type_code}-"
    pattern = re.compile(re.escape(prefix) + r"(\d+)$")

    # Find max existing sequence for this year + type
    max_seq = 0
    for r in rows:
        pn = _s(r.get("permit_number"))
        if not pn:
            continue
        m = pattern.match(pn)
        if m:
            try:
                seq = int(m.group(1))
                if seq > max_seq:
                    max_seq = seq
            except ValueError:
                pass

    col_letter = _col_letter(meta["permit_num_col_idx"])
    sheet_name = meta["sheet_name"]
    assigned = 0

    for r in rows:
        pn = _s(r.get("permit_number"))
        if pn:
            continue
        # A completely blank row shouldn't get a number
        if not any(_s(v) for k, v in r.items() if k not in ("_excel_row", "permit_number")):
            continue

        max_seq += 1
        new_number = f"{prefix}{max_seq}"
        excel_row = r["_excel_row"]
        cell_address = f"{col_letter}{excel_row}"

        if _patch_cell(token, source, sheet_name, cell_address, new_number):
            r["permit_number"] = new_number
            assigned += 1
            print(f"     {source['label']}: assigned {new_number} to row {excel_row}")
        else:
            max_seq -= 1

    return assigned


def _patch_cell(token, source, sheet_name, cell_address, value):
    """PATCH a single worksheet cell via Graph API. Returns True on success."""
    from urllib.parse import quote
    sheet_encoded = quote(sheet_name, safe="")
    url = (
        f"{GRAPH}/drives/{PERMITS_DRIVE_ID}/items/{source['item_id']}"
        f"/workbook/worksheets/{sheet_encoded}/range(address='{cell_address}')"
    )
    body = {"values": [[value]]}
    try:
        r = requests.patch(
            url,
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
            },
            json=body,
            timeout=30,
        )
        if r.status_code >= 400:
            print(f"     ! PATCH {cell_address} failed ({r.status_code}): {r.text[:200]}")
            return False
        return True
    except Exception as e:
        print(f"     ! PATCH {cell_address} error: {e}")
        return False


# ══════════════════════════════════════════════════════════════
# TRANSFORM — per-type entry builders
# ══════════════════════════════════════════════════════════════
def _example_row(row):
    ident = _s(row.get("permit_number")).upper()
    return any(ident.startswith(t) for t in EXAMPLE_TOKENS)


def _dates(row, use_authorized_only=False):
    if use_authorized_only:
        start = _date(row.get("authorized_start"))
        end = _date(row.get("authorized_end"))
    else:
        start = _date(row.get("authorized_start")) or _date(row.get("start_date"))
        end = _date(row.get("authorized_end")) or _date(row.get("end_date"))
    if start and end and end < start:
        start, end = end, start
    return start, end


def _attach_geo(entry, row, allow_multipoint=False):
    geo_type = _s(row.get("geo_type")).lower()
    route_coords = _coords(row.get("route_coords"))
    project_pins = _coords(row.get("project_pins")) if allow_multipoint else None
    lat = _float(row.get("lat"))
    lng = _float(row.get("lng"))

    if geo_type == "line" and route_coords and len(route_coords) >= 2:
        entry["geoType"] = "line"
        entry["coordinates"] = route_coords
        mid = route_coords[len(route_coords) // 2]
        entry["lat"], entry["lng"] = mid[0], mid[1]
        return True

    if allow_multipoint and geo_type == "multipoint" and project_pins:
        entry["geoType"] = "multipoint"
        entry["points"] = project_pins
        centroid_lat = sum(p[0] for p in project_pins) / len(project_pins)
        centroid_lng = sum(p[1] for p in project_pins) / len(project_pins)
        entry["lat"], entry["lng"] = round(centroid_lat, 6), round(centroid_lng, 6)
        return True

    if lat is not None and lng is not None:
        entry["geoType"] = "point"
        entry["lat"], entry["lng"] = lat, lng
        return True

    if route_coords and len(route_coords) >= 2:
        entry["geoType"] = "line"
        entry["coordinates"] = route_coords
        mid = route_coords[len(route_coords) // 2]
        entry["lat"], entry["lng"] = mid[0], mid[1]
        return True
    if allow_multipoint and project_pins:
        entry["geoType"] = "multipoint"
        entry["points"] = project_pins
        centroid_lat = sum(p[0] for p in project_pins) / len(project_pins)
        centroid_lng = sum(p[1] for p in project_pins) / len(project_pins)
        entry["lat"], entry["lng"] = round(centroid_lat, 6), round(centroid_lng, 6)
        return True

    return False


def _row_location(row):
    road = _s(row.get("road"))
    addr = _s(row.get("address"))
    if road and addr:
        if road.lower() in addr.lower():
            return addr
        return f"{road} \u2014 {addr}"
    return road or addr or "Location on file with the Town Clerk"


def _rc_location(row):
    return _s(row.get("road")) or "Location on file with the Town Clerk"


def _dw_location(row):
    return _s(row.get("road")) or "Location on file with the Town Clerk"


def _base_entry(row, permit_type):
    return {
        "id": _s(row.get("permit_number")),
        "permitNumber": _s(row.get("permit_number")),
        "status": _s(row.get("map_status")).lower(),
        "jurisdiction": "town",
        "permitType": permit_type,
    }


def _build_row_entry(row, source):
    start, end = _dates(row, use_authorized_only=False)
    if not start or not end:
        return None, "dates"
    entry = _base_entry(row, "row")
    entry["title"] = _s(row.get("title")) or _s(row.get("type")) or "Permitted work"
    entry["org"] = _s(row.get("org")) or "Applicant on file"
    entry["location"] = _row_location(row)
    entry["startDate"] = start
    entry["endDate"] = end
    if _s(row.get("traffic")):
        entry["traffic"] = _s(row.get("traffic"))
    if _s(row.get("public_contact_name")):
        entry["contactName"] = _s(row.get("public_contact_name"))
    if _s(row.get("public_contact_phone")):
        entry["contactPhone"] = _s(row.get("public_contact_phone"))
    if not _attach_geo(entry, row, allow_multipoint=False):
        return None, "geo"
    return entry, None


def _build_rc_entry(row, source):
    start, end = _dates(row, use_authorized_only=False)
    if not start or not end:
        return None, "dates"
    entry = _base_entry(row, "road_construction")
    entry["title"] = _s(row.get("title")) or _s(row.get("type")) or "Road construction project"
    entry["org"] = _s(row.get("org")) or "Applicant on file"
    entry["location"] = _rc_location(row)
    entry["startDate"] = start
    entry["endDate"] = end
    if _s(row.get("traffic")):
        entry["traffic"] = _s(row.get("traffic"))
    if _s(row.get("public_contact_name")):
        entry["contactName"] = _s(row.get("public_contact_name"))
    if _s(row.get("public_contact_phone")):
        entry["contactPhone"] = _s(row.get("public_contact_phone"))
    if not _attach_geo(entry, row, allow_multipoint=True):
        return None, "geo"
    return entry, None


def _build_dw_entry(row, source):
    start, end = _dates(row, use_authorized_only=True)
    if not start or not end:
        return None, "dates"
    entry = _base_entry(row, "driveway")
    entry["title"] = "New driveway installation"
    entry["org"] = "Property owner on file"
    entry["location"] = _dw_location(row)
    entry["startDate"] = start
    entry["endDate"] = end
    if not _attach_geo(entry, row, allow_multipoint=False):
        return None, "geo"
    return entry, None


_BUILDERS = {
    "row": _build_row_entry,
    "driveway": _build_dw_entry,
    "road_construction": _build_rc_entry,
}


# ══════════════════════════════════════════════════════════════
# BUILD permits.json
# ══════════════════════════════════════════════════════════════
def build_permits_data(all_rows_by_source):
    permits = []
    per_source_stats = {}

    for source, rows in all_rows_by_source:
        stats = {"total": len(rows), "published": 0, "status": 0, "dates": 0, "geo": 0, "example": 0}
        builder = _BUILDERS.get(source["type"])
        if not builder:
            print(f"   ! no builder for type '{source['type']}'; skipping")
            continue

        for row in rows:
            if _example_row(row):
                stats["example"] += 1
                continue

            status = _s(row.get("map_status")).lower()
            if status not in PUBLISHED_STATUSES:
                stats["status"] += 1
                continue

            entry, skip_reason = builder(row, source)
            if entry is None:
                stats[skip_reason] = stats.get(skip_reason, 0) + 1
                continue
            permits.append(entry)
            stats["published"] += 1

        per_source_stats[source["label"]] = stats

    order = {"approved": 0, "proposed": 1}
    permits.sort(key=lambda p: (order.get(p["status"], 2), p["startDate"]))

    return {
        "updated": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "permits": permits,
    }, per_source_stats


# ══════════════════════════════════════════════════════════════
# ENTRY POINT — call from sync.py
# ══════════════════════════════════════════════════════════════
def sync_permits(token, write_github_fn):
    """Fetch every tracker, assign permit numbers to new rows, publish permits.json.
       Returns list of (source_config, rows) tuples for downstream (permit_notify)."""
    print("── Permits ──")

    all_rows = []
    for source in PERMIT_SOURCES:
        rows, meta = fetch_source(token, source)
        if rows is None:
            all_rows.append((source, []))
            continue
        print(f"   {source['label']}: {len(rows)} row(s) in tracker")

        assigned = assign_permit_numbers(token, source, rows, meta)
        if assigned:
            print(f"   {source['label']}: {assigned} new permit number(s) assigned")

        all_rows.append((source, rows))

    data, per_source_stats = build_permits_data(all_rows)

    print(f"   {len(data['permits'])} total published")
    for label, s in per_source_stats.items():
        parts = [f"{s['published']} published"]
        if s.get("status"):   parts.append(f"{s['status']} withheld (status)")
        if s.get("dates"):    parts.append(f"{s['dates']} skipped (dates)")
        if s.get("geo"):      parts.append(f"{s['geo']} skipped (no map location)")
        if s.get("example"):  parts.append(f"{s['example']} example row(s)")
        print(f"     {label}: " + ", ".join(parts))

    payload = json.dumps(data, indent=2, ensure_ascii=False)
    write_github_fn("permits.json", payload)
    print("   permits.json written")

    return all_rows


# ══════════════════════════════════════════════════════════════
# STANDALONE
# ══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    import sys
    try:
        from sync import get_token, write_github_file
    except ImportError:
        print("Run from the repo root so sync.py is importable.", file=sys.stderr)
        sys.exit(1)
    sync_permits(get_token(), write_github_file)
